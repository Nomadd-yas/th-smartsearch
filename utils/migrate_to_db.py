"""Миграция данных из xlsx-файлов в единую SQLite базу данных.

Запуск (один раз перед первым стартом сервера):

    python -m utils.migrate_to_db

Что создаёт:
  data/smartsearch.db
    таблица ste       — все СТЕ: название, категория, производитель,
                        характеристики, нормализованный текст для поиска
    таблица contracts — все контракты с индексом по Идентификатор СТЕ

Если уже существует ste_cache_se_v2.db с предобработанным текстом —
повторная лемматизация пропускается (быстрая миграция ~1-2 мин вместо ~10 мин).
"""

from __future__ import annotations

import sqlite3
import sys
import time
from pathlib import Path

import openpyxl

# Работает и как скрипт (uv run utils/migrate_to_db.py),
# и как модуль (python -m utils.migrate_to_db)
_PROJECT_ROOT = Path(__file__).parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DATA_DIR = _PROJECT_ROOT / "data"
STE_XLSX = DATA_DIR / "TenderHack_СТЕ_20260313.xlsx"
CONTRACTS_XLSX = DATA_DIR / "TenderHack_Контракты_20260313.xlsx"
OLD_CACHE_DB = DATA_DIR / "ste_cache_se_v2.db"

from utils.db import DB_PATH

CONTRACT_COLUMNS = [
    (0,  "Наименование закупки"),
    (1,  "Количество"),
    (2,  "Единица измерения"),
    (3,  "Идентификатор контракта"),
    (4,  "Способ закупки"),
    (5,  "Начальная стоимость контракта"),
    (6,  "Стоимость контракта после заключения"),
    (7,  "% снижения"),
    (8,  "Ставка НДС"),
    (9,  "Дата заключения контракта"),
    (10, "ИНН заказчика"),
    (11, "Регион заказчика"),
    (12, "ИНН поставщика"),
    (13, "Регион поставщика"),
    (14, "Идентификатор СТЕ"),
    (15, "Наименование позиции СТЕ"),
    (16, "Цена за единицу"),
]


def _create_schema(conn: sqlite3.Connection) -> None:
    col_defs = ", ".join(f'"{name}" TEXT' for _, name in CONTRACT_COLUMNS)
    conn.executescript(f"""
        CREATE TABLE IF NOT EXISTS ste (
            ste_id       TEXT PRIMARY KEY,
            name         TEXT NOT NULL,
            category     TEXT NOT NULL,
            manufacturer TEXT,
            characteristics TEXT,
            text_norm    TEXT,
            name_norm    TEXT,
            text_lemma   TEXT
        );

        CREATE TABLE IF NOT EXISTS contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {col_defs}
        );

        CREATE INDEX IF NOT EXISTS idx_contracts_ste_id
            ON contracts("Идентификатор СТЕ");
    """)
    conn.commit()


def _load_old_text_cache() -> dict[str, tuple[str, str, str]]:
    """Загружает text_norm/name_norm/text_lemma из старого кеша если он есть."""
    if not OLD_CACHE_DB.exists():
        return {}
    print(f"  Найден старый кеш {OLD_CACHE_DB.name}, переиспользую нормализацию...")
    conn = sqlite3.connect(OLD_CACHE_DB)
    rows = conn.execute(
        "SELECT ste_id, text_norm, name_norm, text_lemma FROM ste"
    ).fetchall()
    conn.close()
    return {r[0]: (r[1], r[2], r[3]) for r in rows}


def migrate_ste(conn: sqlite3.Connection) -> None:
    print("\n── Миграция СТЕ ──────────────────────────────────────────")
    text_cache = _load_old_text_cache()

    need_lemmatize = not text_cache
    if need_lemmatize:
        print("  Старый кеш не найден — вычисляю нормализацию (займёт ~10 мин)...")
    from utils.search_engine import normalize_text, tokenize_and_lemmatize

    t0 = time.time()
    wb = openpyxl.load_workbook(STE_XLSX, read_only=True)
    ws = wb["СТЕ"]

    batch: list[tuple] = []
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        ste_id = str(row[0])
        name = str(row[1] or "").strip()
        category = str(row[2] or "").strip()
        manufacturer = str(row[3] or "").strip() or None
        characteristics = str(row[4] or "").strip() or None

        if ste_id in text_cache:
            text_norm, name_norm, text_lemma = text_cache[ste_id]
        else:
            raw = name + " " + category
            text_norm = normalize_text(raw)
            name_norm = normalize_text(name)
            text_lemma = " ".join(tokenize_and_lemmatize(text_norm))

        batch.append((ste_id, name, category, manufacturer, characteristics,
                      text_norm, name_norm, text_lemma))
        count += 1

        if len(batch) >= 5_000:
            conn.executemany(
                "INSERT OR REPLACE INTO ste VALUES (?,?,?,?,?,?,?,?)", batch
            )
            conn.commit()
            batch.clear()
            if count % 50_000 == 0:
                print(f"  {count:,} СТЕ...")

    if batch:
        conn.executemany(
            "INSERT OR REPLACE INTO ste VALUES (?,?,?,?,?,?,?,?)", batch
        )
        conn.commit()

    wb.close()
    print(f"  Готово: {count:,} СТЕ за {time.time() - t0:.1f}с")


def migrate_contracts(conn: sqlite3.Connection) -> None:
    print("\n── Миграция контрактов ───────────────────────────────────")
    col_placeholders = ", ".join("?" for _ in CONTRACT_COLUMNS)
    col_names = ", ".join(f'"{name}"' for _, name in CONTRACT_COLUMNS)
    sql = f"INSERT INTO contracts ({col_names}) VALUES ({col_placeholders})"

    t0 = time.time()
    wb = openpyxl.load_workbook(CONTRACTS_XLSX, read_only=True)
    ws = wb["Контракты"]

    batch: list[tuple] = []
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[14] is None:  # Идентификатор СТЕ
            continue
        values = tuple(
            str(row[idx]) if row[idx] is not None else None
            for idx, _ in CONTRACT_COLUMNS
        )
        batch.append(values)
        count += 1

        if len(batch) >= 10_000:
            conn.executemany(sql, batch)
            conn.commit()
            batch.clear()
            if count % 200_000 == 0:
                print(f"  {count:,} контрактов...")

    if batch:
        conn.executemany(sql, batch)
        conn.commit()

    wb.close()
    print(f"  Готово: {count:,} контрактов за {time.time() - t0:.1f}с")


def main() -> None:
    if DB_PATH.exists():
        print(f"БД уже существует: {DB_PATH}")
        answer = input("Пересоздать? Все данные будут удалены [y/N]: ").strip().lower()
        if answer != "y":
            print("Отменено.")
            return
        DB_PATH.unlink()

    print(f"Создаю {DB_PATH}...")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    _create_schema(conn)
    migrate_ste(conn)
    migrate_contracts(conn)

    conn.close()
    size_mb = DB_PATH.stat().st_size / 1024 / 1024
    print(f"\n✓ Миграция завершена. Размер БД: {size_mb:.1f} МБ")
    print(f"  Путь: {DB_PATH}")


if __name__ == "__main__":
    main()
